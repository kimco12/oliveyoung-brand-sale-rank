# -*- coding: utf-8 -*-
"""
올리브영 올영세일(브세) TOP100 브랜드 순위 → HTML 리포트 생성기

입력: "올리브영 26년 브세 top100_k.xlsx" (같은 폴더)
출력: "올리브영_브세_순위리포트.html"

엑셀에 새 세일 기간(예: 26년 12월) 블록이 같은 레이아웃으로 추가되면
이 스크립트만 다시 실행하면 HTML이 자동으로 갱신되고, GitHub Pages에도 자동 배포됩니다.

레이아웃 인식 규칙(자동 탐지, 시트 구조가 바뀌어도 대응):
- 셀 값이 정확히 "순위"인 위치 (r, c)를 블록의 기준점으로 삼는다.
- 제목: 같은 열(c)에서 기준행(r) 위쪽 5행 이내에 있는 첫 텍스트 셀.
- 날짜: 기준행 바로 위 행(r-1), 열 c+1 ~ c+7.
- 데이터: 기준행 다음 행(r+1)부터, 열 c=순위, 열 c+1~c+7=요일별 브랜드.
  순위 열 값이 더 이상 정수가 아니면 블록 종료.

GitHub Pages 자동 배포:
- 인증정보는 코드에 넣지 않고 아래 순서로 읽는다.
    1) 환경변수 GITHUB_TOKEN (+ 선택 GITHUB_REPO, GITHUB_OWNER)
    2) 같은 폴더의 github_config_brandrank.json {"token":"...", "repo":"...", "owner":"..."}
    3) (그래도 없으면) 같은 폴더의 github_config.json 의 token 재사용
      (올리브영 랭킹 대시보드 배포용으로 이미 만들어둔 토큰 — 같은 GitHub 계정)
- 토큰을 못 찾으면 배포는 건너뛰고 HTML 생성까지만 수행한다.
"""
import base64
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, date
from pathlib import Path

import openpyxl
import requests

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
INPUT_XLSX = BASE_DIR / "올리브영 26년 브세 top100_k.xlsx"
OUTPUT_HTML = BASE_DIR / "올리브영_브세_순위리포트.html"

WEEKDAYS = ["일", "월", "화", "수", "목", "금", "토"]

GITHUB_API = "https://api.github.com"
BRANDRANK_CONFIG_FILE = BASE_DIR / "github_config_brandrank.json"
LEGACY_CONFIG_FILE = BASE_DIR / "github_config.json"  # 올리브영 랭킹 대시보드용(토큰 재사용)
DEFAULT_REPO = "oliveyoung-brand-sale-rank"
BRANCH = "main"
PATH_IN_REPO = "index.html"


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s or None


def find_title(ws, r, c):
    for rr in range(r - 1, max(r - 7, 0), -1):
        v = ws.cell(row=rr, column=c).value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return f"기간(행{r},열{c})"


def parse_blocks(ws):
    """시트 안에서 '순위' 헤더를 기준으로 세일 기간 블록들을 찾아 파싱한다."""
    blocks = []
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value != "순위":
                continue
            title = find_title(ws, r, c)
            dates = []
            for i in range(1, 8):
                dates.append(to_date_str(ws.cell(row=r - 1, column=c + i).value))
            if not any(dates):
                continue
            days = []
            data_r = r + 1
            while data_r <= max_row:
                rank_val = ws.cell(row=data_r, column=c).value
                if not isinstance(rank_val, (int, float)):
                    break
                rank = int(rank_val)
                for i in range(7):
                    brand_val = ws.cell(row=data_r, column=c + 1 + i).value
                    if brand_val is None or str(brand_val).strip() == "":
                        continue
                    while len(days) <= i:
                        days.append({})
                    days[i][rank] = str(brand_val).strip()
                data_r += 1
            if not days:
                continue
            blocks.append({
                "title": title,
                "dates": dates[: len(days)],
                "weekdays": WEEKDAYS[: len(days)],
                "rank_by_day": days,  # list[dict[rank]=brand]
            })
    return blocks


def block_sort_key(block):
    for d in block["dates"]:
        if d:
            return d
    return "9999"


def build_period(block):
    """rank->brand(요일별) 구조를 brand-중심 구조로 변환."""
    dates = block["dates"]
    rank_by_day = block["rank_by_day"]
    n_days = len(dates)

    brand_rank_by_day = {}
    for day_idx, day_map in enumerate(rank_by_day):
        for rank, brand in day_map.items():
            brand_rank_by_day.setdefault(brand, [None] * n_days)
            brand_rank_by_day[brand][day_idx] = rank

    last_data_idx = max(
        (i for i, d in enumerate(dates) if d and rank_by_day[i]),
        default=-1,
    )

    rows = []
    for brand, ranks in brand_rank_by_day.items():
        # "최신순위"는 최신 데이터 기준일 당일 순위만 사용한다(그 날 top100 밖이면 이탈로 취급).
        latest_rank = ranks[last_data_idx] if last_data_idx >= 0 else None
        rows.append({"brand": brand, "ranks": ranks, "latest_rank": latest_rank})

    rows.sort(key=lambda x: (x["latest_rank"] is None, x["latest_rank"] if x["latest_rank"] is not None else 9999))

    return {
        "title": block["title"],
        "dates": dates,
        "weekdays": block["weekdays"],
        "last_data_idx": last_data_idx,
        "rows": rows,
    }


def load_deploy_config():
    token = os.getenv("GITHUB_TOKEN", "").strip()
    repo = os.getenv("GITHUB_REPO", "").strip()
    owner = os.getenv("GITHUB_OWNER", "").strip()

    cfg = {}
    if BRANDRANK_CONFIG_FILE.exists():
        try:
            cfg = json.loads(BRANDRANK_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            log.error("github_config_brandrank.json 읽기 실패: %s", e)

    token = token or str(cfg.get("token", "")).strip()
    if not token and LEGACY_CONFIG_FILE.exists():
        try:
            legacy = json.loads(LEGACY_CONFIG_FILE.read_text(encoding="utf-8"))
            token = str(legacy.get("token", "")).strip()
            if token:
                log.info("github_config.json(랭킹 대시보드용)의 토큰을 재사용합니다.")
        except Exception as e:
            log.error("github_config.json 읽기 실패: %s", e)

    repo = repo or str(cfg.get("repo", "")).strip() or DEFAULT_REPO
    owner = owner or str(cfg.get("owner", "")).strip()
    return token, repo, owner, cfg


def save_deploy_config(cfg, **updates):
    cfg.update(updates)
    BRANDRANK_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def gh(method, path, token, **kw):
    headers = kw.pop("headers", {})
    headers.update({
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    })
    return requests.request(method, f"{GITHUB_API}{path}", headers=headers, timeout=60, **kw)


def ensure_repo(token, owner, repo, private=False, description="올리브영 올영세일(브세) TOP100 브랜드 순위 리포트 (자동 갱신)"):
    r = gh("GET", f"/repos/{owner}/{repo}", token)
    if r.status_code == 200:
        return True
    if r.status_code == 404:
        log.info("저장소 없음 → 생성(%s): %s/%s", "private" if private else "public", owner, repo)
        c = gh("POST", "/user/repos", token, json={
            "name": repo, "private": private, "auto_init": True,
            "description": description,
        })
        if c.status_code in (200, 201):
            time.sleep(2)
            return True
        log.error("저장소 생성 실패 status=%s: %s", c.status_code, c.text[:300])
        return False
    log.error("저장소 확인 실패 status=%s: %s", r.status_code, r.text[:300])
    return False


def put_file(token, owner, repo, content_text, path=PATH_IN_REPO, commit_message=None):
    sha = None
    g = gh("GET", f"/repos/{owner}/{repo}/contents/{path}?ref={BRANCH}", token)
    if g.status_code == 200:
        sha = g.json().get("sha")
    content_b64 = base64.b64encode(content_text.encode("utf-8")).decode("ascii")
    payload = {
        "message": commit_message or f"update {path} {time.strftime('%Y-%m-%d %H:%M')}",
        "content": content_b64,
        "branch": BRANCH,
    }
    if sha:
        payload["sha"] = sha
    p = gh("PUT", f"/repos/{owner}/{repo}/contents/{path}", token, json=payload)
    if p.status_code in (200, 201):
        return True
    log.error("파일 커밋 실패(%s) status=%s: %s", path, p.status_code, p.text[:300])
    return False


# 저장소에는 완성된 리포트(index.html)와 아래 소스 파일들만 올린다.
# 엑셀 원본, github_config*.json(토큰 포함) 등은 절대 포함하지 않는다.
SOURCE_FILES_TO_PUBLISH = [
    (BASE_DIR / "generate_brand_rank_report.py", "generate_brand_rank_report.py"),
    (BASE_DIR / "_brand_rank_report_template.html", "_brand_rank_report_template.html"),
    (BASE_DIR / "메크로" / "oliveyoung_brand rank wide.py", "oliveyoung_brand rank wide.py"),
    (BASE_DIR / "BRAND_RANK_README.md", "README.md"),
]


def push_source_files(token, owner, repo):
    for local_path, repo_path in SOURCE_FILES_TO_PUBLISH:
        if not local_path.exists():
            log.warning("소스 파일 없음, 건너뜀: %s", local_path)
            continue
        ok = put_file(
            token, owner, repo,
            local_path.read_text(encoding="utf-8"),
            path=repo_path,
            commit_message=f"update {repo_path} {time.strftime('%Y-%m-%d %H:%M')}",
        )
        if ok:
            log.info("소스 커밋 완료: %s", repo_path)


def ensure_pages(token, owner, repo):
    r = gh("GET", f"/repos/{owner}/{repo}/pages", token)
    if r.status_code == 200:
        return r.json().get("html_url", "")
    c = gh("POST", f"/repos/{owner}/{repo}/pages", token,
           json={"source": {"branch": BRANCH, "path": "/"}})
    if c.status_code in (201, 200):
        return c.json().get("html_url", "")
    if c.status_code == 409:
        r2 = gh("GET", f"/repos/{owner}/{repo}/pages", token)
        return r2.json().get("html_url", "") if r2.status_code == 200 else ""
    log.warning("Pages 활성화 응답 status=%s: %s", c.status_code, c.text[:200])
    return ""


def deploy_to_github_pages(html_text):
    token, repo, owner, cfg = load_deploy_config()
    if not token:
        log.warning("GitHub 토큰 없음 → 배포 건너뜀. github_config_brandrank.json 의 token 을 채우세요.")
        return "skip"

    if not owner:
        u = gh("GET", "/user", token)
        if u.status_code != 200:
            log.error("GitHub 사용자 조회 실패 status=%s: %s", u.status_code, u.text[:200])
            return False
        owner = u.json().get("login", "")
        save_deploy_config(cfg, token=token, repo=repo, owner=owner)
        log.info("GitHub 사용자: %s", owner)

    if not ensure_repo(token, owner, repo, private=False):
        return False
    if not put_file(token, owner, repo, html_text):
        return False
    log.info("index.html 커밋 완료 (%s/%s)", owner, repo)
    push_source_files(token, owner, repo)

    url = ensure_pages(token, owner, repo)
    pages_url = url or f"https://{owner}.github.io/{repo}/"
    log.info("배포 완료 → %s (반영까지 최대 1~2분)", pages_url)
    save_deploy_config(cfg, token=token, repo=repo, owner=owner, pages_url=pages_url)
    return True


def main():
    if not INPUT_XLSX.exists():
        raise SystemExit(f"입력 파일을 찾을 수 없습니다: {INPUT_XLSX}")

    log.info("엑셀 로딩: %s", INPUT_XLSX.name)
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)

    all_blocks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        blocks = parse_blocks(ws)
        log.info("시트 '%s' → %d개 기간 블록", sheet_name, len(blocks))
        all_blocks.extend(blocks)

    if not all_blocks:
        raise SystemExit("기간 블록을 찾지 못했습니다. 엑셀 레이아웃을 확인하세요.")

    all_blocks.sort(key=block_sort_key)
    periods = [build_period(b) for b in all_blocks]

    for p in periods:
        log.info("  - %s: 브랜드 %d개, 날짜 %s", p["title"], len(p["rows"]), [d for d in p["dates"] if d])

    template_path = BASE_DIR / "_brand_rank_report_template.html"
    html_template = template_path.read_text(encoding="utf-8")
    html = html_template.replace(
        "/*__PERIODS_JSON__*/",
        json.dumps(periods, ensure_ascii=False),
    )

    OUTPUT_HTML.write_text(html, encoding="utf-8")
    log.info("생성 완료: %s", OUTPUT_HTML)

    deploy_to_github_pages(html)


if __name__ == "__main__":
    main()
